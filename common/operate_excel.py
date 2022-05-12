import openpyxl
from openpyxl.styles import colors, Font


class OperateExcel(object):

    def __init__(self, filename, sheet_name):
        self.filename = filename
        self.sheet_name = sheet_name

    def open_excel(self):
        self.wb = openpyxl.load_workbook(self.filename)
        self.sh = self.wb[self.sheet_name]

    def write_data(self, row, column, value):
        """写入数据"""
        # 打开文件
        self.open_excel()
        # 设置字体颜色
        color = '006100'
        font = Font(color=colors.BLUE)
        # 写入数据
        # self.sh.cell(row=row, column=column, value=value)
        self.sh.cell(row=row, column=column, value=value).font = font  # 序列
        # 保存文件
        self.wb.save(self.filename)
